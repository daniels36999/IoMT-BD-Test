import tkinter
from PIL import ImageTk, Image
from tkinter import Tk, Label, Button,Entry, Frame, END , messagebox ,ttk
from tkinter import * 
import cv2
import os
import imutils
import datetime
import time
import numpy as np
import pandas as pd
import openpyxl
import serial
import datetime
from openpyxl import load_workbook
from openpyxl.chart import Reference,  LineChart
from git import Repo
import PIL.Image
import PIL.ImageTk
import paho.mqtt.client as mqtt

# formato = '%c'
# ahora = time.strftime(formato)

# fechaActual= datetime.datetime.now()




dato8 = 0
dato9 = 0

def inicio():
    contador=0
    serialArduino = serial.Serial('/dev/ttyACM0',9600)
    while True:
        
        
        cad=serialArduino.readline().decode('ascii')
        datoss=cad.splitlines()
        d0=str(datoss[0])
        d1=d0.replace("b","")
        d2=d1.replace("'","")
        d3=d2.split(",")
        dato1=str(d3[0])
        dato2=str(d3[1])
        dato3=str(d3[2])
        dato4=str(d3[3])
        dato5=str(d3[4])
                    
        dato9=float(dato3)
        if(dato9<1):
            dato9=1
                        
        dato6=float(dato2)/(float(dato9)*float(dato9))
        dato7=round(dato6,2)
        dato8=str(dato7)
                    
        print(dato1,dato2,dato3,dato8,dato5)
        print("--------------------")
        print(contador)
        print("--------------------")
        client = mqtt.Client()
        client.connect("test.mosquitto.org",1883,60)
        client.publish("Cabina/IoMT/Datos", str(dato1)+';'+str(dato2)+';'+str(dato3)+';'+str(dato5)+';'+str(dato5));
        client.disconnect();
        contador=contador+1
        time.sleep(0.5)
        if contador>= 15:
            messagebox.showinfo('IoMT','IoMT realizado exitosamente')
            serialArduino.close()
            break

def ventanaregis():


    
    ventana.withdraw()
    ventana2 = tkinter.Toplevel()
#     ventana2.geometry('600x400')
    ventana2.attributes('-fullscreen', True)
    ventana2.title('REGISTRO')
    labelregistro = Label(ventana2, image=photo1).place(x=0,y=0,relwidth=1.0,relheight=1.0)

#     labelregistro.pack()

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #     

   
#########################################################################################
#########################################################################################
    entradanombre= tkinter.Entry(ventana2, text= 'Nombre',font=('ALGERIAN 20 bold'))
#     entradanombre.grid(row= 1, column = 1)
    entradanombre.place(x=456 , y=207 , width=652, height=70)

    entradaapellido= tkinter.Entry(ventana2, text= 'Apellido',font=('ALGERIAN 20 bold'))
#     entradaapellido.grid(row= 2, column = 1)
    entradaapellido.place(x=456 , y=291 , width=652, height=70)

    entradacedula= tkinter.Entry(ventana2, text= 'Cédula',font=('ALGERIAN 20 bold'))
#     entradacedula.grid(row= 3, column = 1)
    entradacedula.place(x=456 , y=376 , width=652, height=70)
    
    entradacurso = ttk.Combobox(ventana2, state = 'readonly',font=('ALGERIAN 20 bold'))
    entradacurso['values']=['Seleccione','Inicial 1','Inicial 2','Primero','Segundo','Tercero','Cuarto','Quinto','Sexto','Séptimo','Octavo','Noveno','Décimo','Primero de Bachillerato','Segundo de Bachillerato','Tercero de Bachillerato','Básica Acelerado','Docente','Otros']
    entradacurso.current(0)
#     entradacurso.grid(row= 4, column = 1 )
    entradacurso.place(x=456 , y=462 , width=652, height=70)

    
#     entradacurso= tkinter.Entry(ventana2, text= 'Curso')
#     entradacurso.grid(row= 4, column = 1)
    

    def entrenar():
        
        dataPath = '/home/pi/DispositivoFinalIoMT/Data' #Cambia a la ruta donde hayas almacenado Data
        peopleList = os.listdir(dataPath)
        print('Lista de personas: ', peopleList)

        labels = []
        facesData = []
        label = 0

        for nameDir in peopleList:
            personPath = dataPath + '/' + nameDir
            print('Leyendo las imágenes')

            for fileName in os.listdir(personPath):
                print('Rostros: ', nameDir + '/' + fileName)
                labels.append(label)
                facesData.append(cv2.imread(personPath+'/'+fileName,0))

            label = label + 1

        face_recognizer = cv2.face.EigenFaceRecognizer_create()

        print("Entrenando...")
        face_recognizer.train(facesData, np.array(labels))


        face_recognizer.write('modeloEigenFacepruebainterface.xml')

        print("Modelo almacenado...")

    def mensajeregistro():
        messagebox.showinfo('Registro','Registro Realizado Correctamente')
    def mensajeerror():
        messagebox.showinfo('Error','El usuario que desea Registrar ya existe en la base de datos')
    
    def atras():
        ventana2.withdraw()
        ventana.deiconify()
        
    def capturarrostro():
        nombre = entradanombre.get()
        apellido = entradaapellido.get()
        cedula = entradacedula.get()
        curso = entradacurso.get()
        
        if(entradacurso.get()== 'Seleccione'):
            curso = 'Otros'

        personName = nombre + ' ' + apellido
        dataPath = '/home/pi/DispositivoFinalIoMT/Data' #Cambia a la ruta donde hayas almacenado Data
        personPath = dataPath + '/' + personName
        
        
        
        

        if (os.path.isfile('/home/pi/DispositivoFinalIoMT/Todos los Datos/'+ personName +'.csv')):
            mensajeerror()
        
        else:
            os.path.exists(personPath)
            print('Carpeta creada: ',personPath)
            os.makedirs(personPath)
            
            
            cap = cv2.VideoCapture(0,cv2.CAP_V4L)
            

            faceClassif = cv2.CascadeClassifier(cv2.data.haarcascades+'haarcascade_frontalface_default.xml')
            count = 0

            while True:

                ret, frame = cap.read()
                if ret == False: break
                frame =  imutils.resize(frame, width=640)
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                auxFrame = frame.copy()

                faces = faceClassif.detectMultiScale(gray,1.3,5)

                for (x,y,w,h) in faces:
                    cv2.rectangle(frame, (x,y),(x+w,y+h),(0,255,0),2)
                    rostro = auxFrame[y:y+h,x:x+w]
                    rostro = cv2.resize(rostro,(150,150),interpolation=cv2.INTER_CUBIC)
                    cv2.imwrite(personPath + '/rostro_{}.jpg'.format(count),rostro)
                    count = count + 1
                cv2.imshow('frame',frame)

                k =  cv2.waitKey(1)
                if k == 27 or count >= 20:
                
                    break

            cap.release()
            cv2.destroyAllWindows()
            
            entrenar()
#########################################################################################            
            archivo = open ('/home/pi/DispositivoFinalIoMT/Todos los Datos/'+ personName +'.csv','a')
            archivo.write('Nombre '+ ','+ personName + '\n' + 'Cedula' + ',' + cedula +'\n'+ 'Curso' + ',' + curso +'\n')
            archivo.close()
#########################################################################################            
            archivo1 = open ('/home/pi/GitHub-IoMT/IoMT-BD/Datos Almacenados/'+curso+'/'+ personName +'.html','a')
            archivo1.write(
            """
            <!doctype html>
            <html lang="es">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport"
                      content="width=device-width, user-scalable=no, initial-scale=1.0, maximum-scale=1.0, minimum-scale=1.0">
                <meta http-equiv="X-UA-Compatible" content="ie=edge">
                
                <title>HISTORIAL IoMT</title>

                <!-- LIBRERIAS -->
                <script src="https://unpkg.com/xlsx@0.16.9/dist/xlsx.full.min.js"></script>
                <script src="https://unpkg.com/file-saverjs@latest/FileSaver.min.js"></script>
                <script src="https://unpkg.com/tableexport@latest/dist/js/tableexport.min.js"></script>
                <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.0.0-beta2/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-BmbxuPwQa2lc/FVzBcNJ7UAyJxM6wuqIj61tLrc4wSX0szH/Ev+nYRRuWlolflfl" crossorigin="anonymous">

            </head>

            <body>
            <br><br>
            <div class="container">
                <div class="card">
                    <div class="card-header">
                        REPORTE IOMT
                    </div>
                    <div class="card-body">
                        <button id="btnExportar" class="btn btn-success">
                            <i class="fas fa-file-excel"></i> DESCARGAR HISTORIAL
                        </button>

                        <table id="tabla" class="table table-border table-hover">
                            <thead>
                            <tr><th>Nombre:</th><th>"""+nombre+"""</th></tr>				
                            <tr><th>Apellido:</th><th>"""+apellido+"""</th></tr>
                            <tr><th>Cedula:</th><th>"""+cedula+"""</th></tr>
                            <tr><th>Curso:</th><th>"""+curso+"""</th></tr>
                            <tr><th></th><th></th></tr>				
                            
                            <tr><th>Fecha</th>
                                <th>Hora</th>
                                <th>Temperatura</th>
                                <th>Peso</th>
                                <th>Altura</th>
                                <th>IMC</th>
                                <th>O2Sat</th>
                            </tr>
                            </thead>
                            <tbody>

                        <!-- AGREGAR TABLA-->
                            
                            
                            </tbody>
                        </table>
                    </div>
                </div>
            </div>
                <!-- SCRIPT PARA EXPORTAR -->
                <script>
                    const $btnExportar = document.querySelector("#btnExportar"),
                        $tabla = document.querySelector("#tabla");
                    $btnExportar.addEventListener("click", function() {
                        let tableExport = new TableExport($tabla, {
                            exportButtons: false, // No queremos botones
                            filename: "Reporte de prueba", //Nombre del archivo de Excel
                            sheetname: "Reporte de prueba", //Título de la hoja
                        });
                        let datos = tableExport.getExportData();
                        let preferenciasDocumento = datos.tabla.xlsx;
                        tableExport.export2file(preferenciasDocumento.data, preferenciasDocumento.mimeType, preferenciasDocumento.filename, preferenciasDocumento.fileExtension, preferenciasDocumento.merges, preferenciasDocumento.RTL, preferenciasDocumento.sheetname);
                    });
                </script>
            </body>
            </html>
            """)
            archivo1.close()
                        
#########################################################################################
            wb = openpyxl.Workbook()
            ws=wb.active
            temp=(["Nombre", personName])
            temp2=(["Cédula", cedula])
            temp3=(["Curso", curso])
            temp5=(["Fecha","Hora", "Temperatura","Peso","Altura","IMC","O2Sat"])
            ws.append(temp)
            ws.append(temp2)
            ws.append(temp3)
            ws.append(temp5)
            wb.save('/home/pi/DispositivoFinalIoMT/DatosExcel/'+personName+".xlsx")
     
            mensajeregistro()
            
            entradanombre.delete(0,END) 
            entradaapellido.delete(0,END) 
            entradacedula.delete(0,END) 
            entradacurso.delete(0,END) 
        


    botonaceptar = tkinter.Button(ventana2, text = 'Aceptar', command  = capturarrostro, padx=74, pady=18,font=('ALGERIAN 20 bold') )
    botonaceptar.place(x=289, y=606)
    #botonregistro.pack()  // aparece el boton en la ventana
    #botonaceptar.grid(row =6, column=1)
    
    botonregresar = tkinter.Button(ventana2, text = 'Regresar', command  = atras , padx=66, pady=18, font=('ALGERIAN 20 bold'))
    botonregresar.place(x=722, y=606)
    #botonregistro.pack()  // aparece el boton en la ventana
    #botonregresar.grid(row =6 , column=2)
    

def actualizargithub():
    
    repo = Repo('/home/pi/GitHub-IoMT/IoMT-BD')  # if repo is CWD just do '.'
    repo.index.add(['/home/pi/GitHub-IoMT/IoMT-BD/Datos Almacenados'])
    repo.index.commit('actualizando')
    origin = repo.remote('origin')
    origin.push()
    
def reconocimientof():
    
  
    serialArduino = serial.Serial('/dev/ttyACM0',9600)
    
    dataPath = '/home/pi/DispositivoFinalIoMT/Data' #Cambia a la ruta donde hayas almacenado Data
    imagePaths = os.listdir(dataPath)
    print('imagePaths=',imagePaths)

    face_recognizer = cv2.face.EigenFaceRecognizer_create()

    face_recognizer.read('modeloEigenFacepruebainterface.xml')


    cap = cv2.VideoCapture(0,cv2.CAP_V4L)


    faceClassif = cv2.CascadeClassifier(cv2.data.haarcascades+'haarcascade_frontalface_default.xml')
    
    count1 = 0
    
    while True:
        ret,frame = cap.read()
        if ret == False: break
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        auxFrame = gray.copy()

        faces = faceClassif.detectMultiScale(gray,1.3,5)
        

        for (x,y,w,h) in faces:
            rostro = auxFrame[y:y+h,x:x+w]
            rostro = cv2.resize(rostro,(150,150),interpolation= cv2.INTER_CUBIC)
            result = face_recognizer.predict(rostro)

            cv2.putText(frame,'{}'.format(result),(x,y-5),1,1.3,(255,255,0),1,cv2.LINE_AA)

            if result[1] < 4500:
                cv2.putText(frame,'{}'.format(imagePaths[result[0]]),(x,y-25),2,1.1,(0,255,0),1,cv2.LINE_AA)
                cv2.rectangle(frame, (x,y),(x+w,y+h),(0,255,0),2)
                nombrefinal = '{}'.format(imagePaths[result[0]])
                nombreparallenar = str(nombrefinal)
                count1 = count1 + 1

            else:
                cv2.putText(frame,'Desconocido',(x,y-20),2,0.8,(0,0,255),1,cv2.LINE_AA)
                cv2.rectangle(frame, (x,y),(x+w,y+h),(0,0,255),2)
                
        cv2.imshow('frame',frame)
        
        print(count1)
        
        k =  cv2.waitKey(1)
        if k == 27 or count1 >= 10:
        
            break
        

    
    cap.release()
    cv2.destroyAllWindows()
       

    


 # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
 
    datos=pd.read_csv('/home/pi/DispositivoFinalIoMT/Todos los Datos/'+nombreparallenar+'.csv', header= None)
    df=pd.DataFrame(datos)
    a = str(df.iat[2,1])
    print(a)
    
#     serialArduino = serial.Serial('/dev/ttyACM0',9600)
    cont1=0
    cont2=0
    
#    serialArduino = serial.Serial('/dev/ttyACM0',9600)
    
    while True :
#         cad=serialArduino.readline().decode('ascii')
        
#         serialArduino = serial.Serial('/dev/ttyACM0',9600)
        cad=serialArduino.readline().decode('ascii')
        if(cont1>=10):
            
            
            cad=serialArduino.readline().decode('ascii')
            datoss=cad.splitlines()
            d0=str(datoss[0])
            d1=d0.replace("b","")
            d2=d1.replace("'","")
            d3=d2.split(",")
            dato1=str(d3[0])
            dato2=str(d3[1])
            dato3=str(d3[2])
            dato4=str(d3[3])
            dato5=str(d3[4])
            
            dato9=float(dato3)
            if(dato9<1):
                dato9=1
                
            dato6=float(dato2)/(float(dato9)*float(dato9))
            dato7=round(dato6,2)
            dato8=str(dato7)
            
            print(dato1,dato2,dato3,dato8,dato5)
            print("--------------------")
            cont1=cont1+1
            ahora=datetime.datetime.now()
            ahora1=ahora.strftime("%d/%m/%y")
            ahora2=ahora.strftime("%Hh/%Mm/%Ss")
            client = mqtt.Client()
            client.connect("test.mosquitto.org",1883,60)
            client.publish("Cabina/IoMT/Datos", str(dato1)+';'+str(dato2)+';'+str(dato3)+';'+str(dato5)+';'+str(dato5));
            client.disconnect();
            
 
        if(cont2==20):
######################################################################################################
            with open('/home/pi/GitHub-IoMT/IoMT-BD/Datos Almacenados/'+a+'/'+nombreparallenar+'.html',"r") as f:
                newline=[]
                for word in f.readlines():        
                    newline.append(word.replace("<!-- AGREGAR TABLA-->","""<tr><td>"""+ahora1+"""</td><td>"""+ahora2+"""</td><td>"""+dato1+"""</td><td>"""+dato2+"""</td><td>"""+dato3+"""</td><td>"""+str(dato8)+"""</td><td>"""+dato5+"""</td></tr>
                            <!-- AGREGAR TABLA-->"""))  ## Replace the keyword while you copy.  

            with open('/home/pi/GitHub-IoMT/IoMT-BD/Datos Almacenados/'+a+'/'+nombreparallenar+'.html',"w") as f:
                for line in newline:
                    f.writelines(line)
                f.close()
            

######################################################################################################
            wb2=openpyxl.load_workbook('/home/pi/DispositivoFinalIoMT/DatosExcel/'+nombreparallenar+'.xlsx')
            ws=wb2.active
            temp=([ahora1, ahora2, dato1,dato2,dato3,str(dato8),dato5])
            ws.append(temp)
            wb2.save('/home/pi/DispositivoFinalIoMT/DatosExcel/'+nombreparallenar+'.xlsx')
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 
            client = mqtt.Client()
            client.connect("test.mosquitto.org",1883,60)
            client.publish("Cabina/IoMT/Datos", str(dato1)+';'+str(dato2)+';'+str(dato3)+';'+str(dato8)+';'+str(dato5));
            client.disconnect();
#            serialArduino.close()
    
            cont1=0
            cont2=0
            
            actualizargithub()
#             serialArduino.close()
            break

        cont1=cont1+1
        cont2=cont2+1
        print(cont1,cont2)
        time.sleep(0.5)
    


 # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #    


    

    
    messagebox.showinfo('Datos','Datos Almacenados Correctamente')
    client = mqtt.Client()
    client.connect("test.mosquitto.org",1883,60)
    client.publish("Cabina/IoMT/Datos", str("00:00")+';'+str("00:00")+';'+str("00:00")+';'+str("00:00")+';'+str("00:00"));
    client.disconnect();
    
#     serialArduino.close()

#     ventana3.withdraw()
#     botonsalir = tkinter.Button(ventana, text = 'Autentificacion', command = reconocimientof)
    ventana.deiconify() 




ventana = tkinter.Tk()
# ventana.geometry('600x600')  #TAma;o de la ventana principal
ventana.title('Ventana Principal')
ventana.attributes('-fullscreen', True)

im = PIL.Image.open("/home/pi/DispositivoFinalIoMT/fondoprincipal.png")
im1 = PIL.Image.open("/home/pi/DispositivoFinalIoMT/fondoregistro.png")
photo = PIL.ImageTk.PhotoImage(im)
photo1 = PIL.ImageTk.PhotoImage(im1)
label = Label(ventana, image=photo).place(x=0,y=0,relwidth=1.0,relheight=1.0)


# actualizargithub()

# etiqueta =  tkinter.Label(ventana, text= 'Bienvenido', bg = 'gray',  height=5 , width=40)  #Etiqueta de Entrada 
# #etiqueta.grid(row = 0, column = 4) # Aparece la Linea Gris en toda la parte superior
# etiqueta.place(x=450 , y=20)
botonregistro = tkinter.Button(ventana, text = 'Registro', command  = ventanaregis , padx=74, pady=56,font=('ALGERIAN 20 bold'))
#botonregistro.pack()  // aparece el boton en la ventana
#botonregistro.grid(row =2, column=2)
botonregistro.place(x=203 , y=390)


botonautentificacion = tkinter.Button(ventana, text = 'Autentificación', command = reconocimientof,padx=25, pady=56,font=('ALGERIAN 20 bold'))
#botonautentificacion.pack()
#botonautentificacion.grid(row =1, column=2)
botonautentificacion.place(x=790, y=390)

####################################################################################################
botoniomt = tkinter.Button(ventana, text = 'IOMT \n usuarios no registrados', command = inicio ,padx=25, pady=50,font=('ALGERIAN 15 bold'))
#botonautentificacion.pack()
#botonautentificacion.grid(row =1, column=2)
botoniomt.place(x=480, y=570)

########################################################################################

serialArduino = serial.Serial('/dev/ttyACM0',9600)




ventana.mainloop()


